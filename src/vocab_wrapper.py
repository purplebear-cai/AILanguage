import openpyxl
import streamlit as st

# Load vocabulary from Excel file at the start
try:
    workbook = openpyxl.load_workbook("data/vocabulary/vocabulary.xlsx")
    sheet = workbook.active
    vocabulary = {}
    for row in sheet.iter_rows(min_row=2):  # Start from row 2 (skip headers)
        english = row[0].value
        romaji = row[1].value
        kanji = row[2].value
        examples = row[3].value.split(", ")  # Split example sentences
        vocabulary[english] = {"romaji": romaji, "kanji": kanji, "examples": examples}
except FileNotFoundError:
    vocabulary = {}
    st.warning("Vocabulary file not found. Starting with an empty dictionary.")

def add_to_dictionary(english, romaji, kanji, example_sentences, filepath):
    """Adds a vocabulary word to an Excel file if it doesn't already exist."""
    if english in vocabulary:
        st.warning(f"Word '{english}' already exists in the dictionary.")
        return

    try:
        workbook = openpyxl.load_workbook(filepath)
        sheet = workbook.active
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        # Add headers
        sheet['A1'] = "English"
        sheet['B1'] = "Romaji"
        sheet['C1'] = "Kanji"
        sheet['D1'] = "Example Sentences"

    # Add the new word to the next available row
    row = sheet.max_row + 1
    sheet.cell(row=row, column=1).value = english
    sheet.cell(row=row, column=2).value = romaji
    sheet.cell(row=row, column=3).value = kanji
    sheet.cell(row=row, column=4).value = ", ".join(example_sentences)  # Join sentences with commas

    # Save the workbook
    workbook.save(filepath)
    vocabulary[english] = {"romaji": romaji, "kanji": kanji, "examples": example_sentences}  # Update in-memory dictionary
    st.success(f"Word added to {filepath}")